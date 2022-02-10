from re import sub
import subprocess
import os
from timeit import timeit
from openpyxl import load_workbook

def setup_worksheet(filename):
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    current_col = sheet.min_column
    current_row = sheet.max_row + 1

    prev_nx = sheet.cell(row=current_row - 1, column = current_col).value
    if prev_nx == "nx":
        nx = 129
    else:
        nx = 2 * (prev_nx - 1) + 1

    sheet.cell(row=current_row, column=current_col).value = nx

    return workbook, sheet, current_row, current_col + 1

def save_to_worksheet(worksheet, row, col, result):
    worksheet.cell(row=row, column=col).value = result

filename = "../../execution_timings.xlsx"
workbook, timings_sheet, excel_row, next_excel_col = setup_worksheet(filename)

compiler_cmds = [
    "gfortran -O3",
    "intel_compile.bat"
    ]
reps = 5

for cmd in compiler_cmds:
    if cmd.endswith(".bat"):
        compile_cmd = cmd
    else:
        compile_cmd = cmd + " -o output.exe 2D_compressible.f90"

    print("\nCompiling using: " + compile_cmd)
    subprocess.call(compile_cmd, env=os.environ, shell=True, 
                    stdout=subprocess.DEVNULL, stderr=subprocess.STDOUT)
    print("Successfully compiled. Running output.exe")

    elapsed_time = timeit(
        stmt = "subprocess.run('output.exe', stdout=subprocess.DEVNULL)",
        setup = "import subprocess",
        number = reps) / reps

    print(f"Saving elapsed time {elapsed_time:3f} to worksheet")
    save_to_worksheet(timings_sheet, excel_row, next_excel_col, elapsed_time)
    next_excel_col += 1

workbook.save(filename)
