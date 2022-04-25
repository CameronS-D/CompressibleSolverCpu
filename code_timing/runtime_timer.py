import subprocess
import os
from timeit import timeit
import openpyxl
from datetime import datetime

def setup_worksheet(filename):
    if os.path.exists(filename):
        workbook = openpyxl.load_workbook(filename)
    else:
        workbook = openpyxl.Workbook()

    sheet = workbook.active
    return workbook, sheet

def save_to_worksheet(worksheet, col, result, row=None):
    if not row:
        letter_col = openpyxl.utils.get_column_letter(col)
        col_cells = [cell for cell in worksheet[letter_col]]

        # Find last cell in column with a value
        row = worksheet.max_row
        for cell in col_cells[::-1]:
            if cell.value is not None:
                break
            row -= 1
        # We need the cell after which is empty
        row += 1

    worksheet.cell(row=row, column=col).value = result

def change_nx_value(filename, old_vals, new_val):
    if filename.endswith(".cu"):
        two_before_nx_line = 12
    else:
        two_before_nx_line = 9

    # Read in full file
    with open(filename, "r") as f:
        firstlines = ""
        for line_idx, line in enumerate(f):
            firstlines += line
            if line_idx == two_before_nx_line:
                break

        nx_line = f.readline()
        rest_of_file = f.read()

    # Change nx value on single line
    for old_val in old_vals:
        nx_line = nx_line.replace(str(old_val), str(new_val), 1)

    full_file = firstlines + nx_line + rest_of_file

    # Write out file
    with open(filename, "w") as f:
        f.write(full_file)

def is_output_correct(correct_output, test_code_cmd):
    print(f"Testing code output. Calling test code: {test_code_cmd}")

    proc = subprocess.run(test_code_cmd, stdout=subprocess.PIPE, encoding="UTF-8")
    test_output = proc.stdout

    print("Comparing output...")
    cor_output_list = correct_output.split()
    test_output_list = test_output.split()
    if len(cor_output_list) != len(test_output_list):
        print("Error: Test code output is too short.")
        print(f"Correct length: {len(cor_output_list)}, Test length: {len(test_output_list)}")
        return False 

    for cor_line, test_line in zip(cor_output_list, test_output_list):
        correct = cor_line[:10]
        test = test_line[:10]

        if correct == test:
            continue
        try:
            if round(float(correct) * 1e5) == round(float(test) * 1e5):
                continue
        except ValueError:
            # Tried to convert word to float - output must be wrong
            pass
        
        print("Error: Test code value is incorrect.")
        print(f"Correct value: {correct}")
        print(f"Output value: {test}")
        return False
    return True

def get_correct_output(filename):
    # Compile
    original_code_cmd = "nvfortran -fast -O3 -o correct_output.exe " + filename
    subprocess.run(original_code_cmd.split(), stdout=subprocess.DEVNULL, stderr=subprocess.STDOUT)
    # Run output
    proc = subprocess.run("./correct_output.exe", stdout=subprocess.PIPE, stderr=subprocess.STDOUT, encoding="UTF-8")
    # Tidy up
    os.remove("./correct_output.exe")

    return proc.stdout

if os.path.isdir("code_timing"):
    os.chdir("code_timing")

f90_file = os.path.join("..", "2D_compressible.f90")
if not os.path.exists(f90_file):
    print(f"Error: Fortran file not found: {f90_file}")
    quit(1)

original_f90_file = "original_2D_compressible.f90"
if not original_f90_file or not os.path.exists(original_f90_file):
    print("Warning: Original code not found")
    original_f90_file = None

cuda_file = os.path.join("CompressibleSolverCUDA", "CompressibleSolverCUDA", "kernel.cu")
if not cuda_file or not os.path.exists(cuda_file):
    print("Warning: CUDA code not found")
    cuda_file = None


excel_file = os.path.join("execution_timings.xlsx")
workbook, timings_sheet = setup_worksheet(excel_file)

compiler_cmds = [
    # ["gfortran -O3 -march=native", "gfortran"],
    # ["gfortran -O3 -march=native -fopenacc -foffload=nvptx-none", "gfortran gpu"],
    # ["ifort -O3 -xhost", "ifort"],
    # ["ifort -O3 -xhost -qopenmp", "ifort parallel"],
    # ["nvfortran -fast -O3", "nvidia"],
    # ["nvfortran -fast -O3 -stdpar=multicore", "nvidia parallel"],
    # ["nvfortran -fast -O3 -stdpar=gpu -gpu=cc70", "nvidia gpu"],
    # ["nvcc -arch=sm_70 -Xptxas -O3", "CUDA"]
    ["nvfortran -fast -O3 -acc=gpu -gpu=cc61,cuda11.6", "openacc"]
]

mesh_nx_options = [129, 257, 513, 1025, 2049]
reps = 3

# Add current date/time and seperators to excel sheet
time_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
min_col = timings_sheet.min_column
next_row = timings_sheet.max_row+1
save_to_worksheet(timings_sheet, min_col, time_str, row=next_row)
for col in range(min_col + 1, min_col + len(compiler_cmds) + 2):
    save_to_worksheet(timings_sheet, col, "-", row=next_row)

for val in ["nx"] + mesh_nx_options:
    save_to_worksheet(timings_sheet, min_col, val)

if original_f90_file:
    correct_outputs = {}
    for nx in mesh_nx_options:
        print(f"Calculating correct output for nx = {nx}")
        change_nx_value(original_f90_file, mesh_nx_options, nx)
        output = get_correct_output(original_f90_file)
        correct_outputs[nx] = output
    
    change_nx_value(original_f90_file, mesh_nx_options, mesh_nx_options[0])
    print("Finshed getting correct outputs")

next_excel_col = timings_sheet.min_column
for cmd, option_name in compiler_cmds:
    if cmd.startswith("nvcc"):
        if not cuda_file:
            print(f"CUDA file not found. Skipping compilation with {cmd}")
            continue

        code_file = cuda_file
    else:
        code_file = f90_file

    compile_cmd = cmd + " -o output.exe " + code_file
    next_excel_col += 1

    for nx_option in mesh_nx_options:
        # Change value of nx in fortran code and add new nx row to spreadsheet
        print(f"\nChanging nx value to {nx_option}")
        change_nx_value(code_file, mesh_nx_options, nx_option)

        if nx_option == mesh_nx_options[0]:
            save_to_worksheet(timings_sheet, next_excel_col, option_name)

        print("\nCompiling using: " + compile_cmd)
        subprocess.run(compile_cmd.split(), stdout=subprocess.DEVNULL, stderr=subprocess.STDOUT)

        if os.path.exists("output.exe"):
            print("Successfully compiled. Running output.exe")
        else:
            print("Compile failed. Moving to next compiler command.")
            break

        if original_f90_file:
            if is_output_correct(correct_outputs[nx_option], "./output.exe"):
                print("Code output is CORRECT. Measuring runtime.")
            else:
                print("INCORRECT output file produced. Moving to next compiler command.")
                save_to_worksheet(timings_sheet, next_excel_col, "N/A")
                os.remove("output.exe")
                break

        elapsed_time = timeit(
            stmt = "subprocess.run('./output.exe', stdout=subprocess.DEVNULL)",
            setup = "import subprocess",
            number = reps) / reps

        print(f"Saving elapsed time {elapsed_time:3f} to worksheet")
        save_to_worksheet(timings_sheet, next_excel_col, elapsed_time)

        os.remove("output.exe")
        workbook.save(excel_file)

    # Change back to original value for next command
    change_nx_value(code_file, mesh_nx_options, mesh_nx_options[0])

print(f"\nCompleted test run. Saving worksheet to {excel_file}.")
workbook.save(excel_file)
